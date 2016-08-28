import os
import math
import string
import openpyxl
import time
import tkinter.filedialog

###############################################################################


class Docket():
    # This class is the structure of the docketrequirements object.
    # Docket.name is a string containing the name of the docket.
    # Docket.requirements is a list of all cell codes that the docket includes

    def __init__(self, name):
        self.name = name
        self.requirements = []


###############################################################################
#   programStart() is the main function directly called by the GUI.           #
#                                                                             #
#   This function only contains error handling and reporting. The functions   #
#   that directly read log/template files are listed below this one.          #
###############################################################################

# The reason why this function takes the MainFrame object as a parameter is
# because I did not know how else to access the changeFrame function located in
# the MainWindow object. I will consider different error reporting methods in
# the future so that a function like this would take just the
# MainFrame.inputlist and MainFrame.templateinput.selectedfilename variables
# as parameters as it should.

def programStart(self):

    def msg(input):
        # Make the message change function easier to access.
        self.parent.workingmessage.set(input)

    def frame(input):
        # Make the changeFrame function easier to access.
        self.parent.changeFrame(input)

    # Set debug to true to ignore file checks (because you have a preset list)
    # Set logdebug to true to log docket and cellcount objects to console.
    debug = False
    logdebug = False

    msg("Please wait, the application is running.")
    frame("MSG")

    template = ""
    template = self.templateinput.selectedfilename.get()
    inputslist = []
    inputslist = list(self.inputlist)

    # Doubles of the variables are listed above for easy debugging.
    # Comment out the second instance of the variable to use the preset
    # filelist for debugging

    # docketrequirements is a list of Docket objects (see above) that each
    # contain a list of cell codes required for their respective jobs.
    docketrequirements = []
    # cellcounts is a list of cellcount dictionaries that contain the
    # quantities of cellcodes mailed in each cycle (log file)
    cellcounts = []
    # filelist is a collection of all files input into the GUI. The first item
    # should always be the template.
    filelist = []

    # Check to see if a template file was included.
    if len(template) < 1 and not debug:
        msg("No template file was selected.")
        frame("BACK")
        return
    elif not debug:
        filelist.append(template)

    # Add all inputs given that are real to the filelist list.
    # If you want to do more validation for inputs, such as ensuring that users
    # have access to the files, this is where it would happen.
    for rawinput in inputslist:
        try:
            if os.path.isfile(rawinput) and not debug:
                item = rawinput
                filelist.append(item)
        except:
            msg("An error has occured when reading the selected files")
            frame("BACK")
            return

    # If debugging is active, use a preset list of files (so you don't have to
    # manually add them). The first file should always be the template file.
    # The rest should all be valid files. Make sure to properly escape the
    # backslashes.
    if debug:
        filelist = []
    if len(filelist) < 2:
        msg("There were no files input.")
        frame("BACK")
        return

    # The first major parsing function. The output of this function is a list
    # of docket objects, each with the form:
    #       docket = {
    #                   'name' = docketnumber
    #                   'cycle' = cyclenumber
    #                   'requirements' = ["list", "of", "cell", "codes"]
    #                }
    try:
        docketrequirements = programParseTemplate(filelist[0])
    except:
        msg("An error has occured when parsing the template file.")
        frame("BACK")
        return
    if not docketrequirements:
        msg("You do not have permission to read " + filelist[0])
        frame("BACK")
        return

    # In case the files processed are very large or numerous, this is a message
    # to list the dockets being processed by the program. This will return all
    # dockets listed (properly) in the docket template. If this screen shows
    # incorrect dockets, then there is a problem with the template file.
    update = ""
    for item in docketrequirements:
        update = update + item.name + ", "
    msg('''Please wait, the application is running.
The following Dockets are being processed (nothing saved yet):

''' + update)

    # programParseLogs is run one at a time for each log file that is entered.
    # This is so that problem files can be correctly error reported and so
    # count information can be separated.
    #
    # The programParseLogs function returns a single dictionary that contains
    # only one cycle's worth of data. The format is as follows:
    #    cellcount = {
    #                   'name' = LOGNAME
    #                   'cellcode1' = QUANTITY
    #                   'cellcode2' = QUANTITY
    #                 }
    for item in filelist[1:]:
        currentfile = str(item)
        try:
            singlelog = programParseLogs(item)
        except:
            msg('''An error has occured when parsing a file. File location:

''' + currentfile)
            frame("BACK")
            return

        if not singlelog:
            msg("You do not have permission to read " + currentfile)
            frame("BACK")
            return
        else:
            cellcounts.append(dict(singlelog))

    # Logging should be done here. This is where data collection ends and file
    # creation begins. Add any other logging functions below if you want to
    # check the ability for the program to properly parse template files.
    if logdebug:
        programDebug(docketrequirements, cellcounts)

    # programCreateExcelFile handles all spreadsheet creation. It
    # takes all the docket information and combines them in a
    # single workbook. It CANNOT update existing sheets and so the
    # information inside should be moved to an archivable file
    try:
        docketfile = programCreateExcelFile(docketrequirements, cellcounts)
    except:
        msg("An error has occured when creating the excel files.")
        frame("BACK")
        return

    # Saving is the last step.
    msg("Please save your Excel Workbook. The default name is \"Generated on YYYY-MM-DD\"")
    try:
        programSaveExcelFile(docketfile)
    except:
        msg("An error has occured while the file was being saved.")
        frame("BACK")
        return

    # Nothing broke!
    msg("Program Complete. Press here to return to the main screen.")
    frame("BACK")
    return

    ##################
    # END OF PROGRAM #
    ##################

###############################################################################
#   safeOpen()                                                                #
#                                                                             #
#   This is a safe open() function that checks for user authentication.       #
#   Right after it is used, there should be something that checks to see if   #
#   the output is False.                                                      #
###############################################################################


def safeOpen(filelocation, opentype):

    # If the user has permission, open as normal #
    if os.access(filelocation, os.R_OK):
        return open(filelocation, opentype)
    else:
        return False


###############################################################################
#   programDebug()                                                            #
#                                                                             #
#   This is a simple debug function that prints all objects in use to the     #
#   Python Console.                                                           #
###############################################################################
def programDebug(docketrequirements, cellcounts):
    print("CELL CODES IN EACH DOCKET:")
    for item in docketrequirements:
        print("\n" + item.name)
        for value in item.requirements:
            print(value)
    print("\n CELL COUNT DATA:")
    for item in cellcounts:
        print("\n" + str(item))


###############################################################################
#   programSaveExcelFile()                                                    #
#                                                                             #
#   This generates file details and handles saving.                           #
###############################################################################
def programSaveExcelFile(docket):

    defaultname = "Generated " + time.strftime("%Y-%m-%d") + " at " + time.strftime("%I%p") + ".xlsx"
    excel = [('Microsoft Excel 2007-2013 XML', '.xlsx')]
    docket.save(tkinter.filedialog.asksaveasfilename(filetypes=excel, initialfile=defaultname))


###############################################################################
#   programParseTemplate()                                                    #
#                                                                             #
#   This function goes through the template file and considers all lines      #
#   beneath a line starting with "Docket" (case-insensitive) as a part of the #
#   docket requirements for that job. This means that the codes in the        #
#   template file are case sensitive and any mistakes during its creation     #
#   will cause this program to completely ignore valid data.                  #
###############################################################################
def programParseTemplate(filelocation):

    rawfile = safeOpen(filelocation, 'r')
    if not rawfile:
        return False

    # incompletedocket holds incomplete Docket objects. When an empty line is
    # reached in the template file, the information in incompletedocket is
    # pushed into a Docket object and appended to the docketrequirement array.
    # Docketrequirements holds all completed Docket objects
    incompletedocket = []
    docketrequirements = []

    # The program ignores all lines unless filldocket is set to True
    # Because we do not know if the first lines list the docket name we
    # initialize it as False
    filldocket = False

    for i, rawline in enumerate(rawfile):
        line = rawline.replace('\n', '')

        # In case someone forgot to use a space between the docket codes from
        # the next docket, push the existing docket information into
        # docketrequirements and make a new object. Don't reset filldocket
        # because we will continue collection immediately after.
        if filldocket and bool(line.split(" ")[0].upper() == "DOCKET"):
            if len(incompletedocket) > 1:  # Make sure it has info in it.
                finisheddocket = incompletedocket[0]
                finisheddocket.requirements = incompletedocket[1:]
                docketrequirements.append(finisheddocket)

            incompletedocket = []
            newdocket = Docket(line)
            incompletedocket.append(newdocket)

        elif filldocket and bool(line != ""):
            incompletedocket.append(line)

        # If there is an empty line and we are currently collecting docket data
        # stop collection, push the current docket information into the Docket
        # object and reset.
        elif filldocket and bool(line == ""):
            if len(incompletedocket) > 1:  # Make sure it has info in it.
                finisheddocket = incompletedocket[0]
                finisheddocket.requirements = incompletedocket[1:]
                docketrequirements.append(finisheddocket)

            incompletedocket = []
            filldocket = not filldocket

        # If we have not started collection yet (because it is the start of a
        # file or we just finished collecting one docket), then wait for a line
        # the starts with the word "DOCKET". If we find one, start collection.
        elif line.split(" ")[0].upper() == "DOCKET":
            newdocket = Docket(line)
            incompletedocket.append(newdocket)
            filldocket = not filldocket

    # Just in case someone didn't end with a blank line, if we have docket
    # information in incompletedocket, then create a new Docket object with
    # that information.
    if len(incompletedocket) > 1:
        finisheddocket = incompletedocket[0]
        finisheddocket.requirements = incompletedocket[1:]
        docketrequirements.append(finisheddocket)

    return docketrequirements


###############################################################################
#   programParseLogs()                                                        #
#                                                                             #
#   This function returns a list of dictionaries containing cell code counts. #
###############################################################################
def programParseLogs(filelocation):

    log = {}
    filename = os.path.basename(filelocation)

    # This Log dictionary contains a list of all cellcodes and quantities found
    # in the input log file. Cellcodes are keys, quantities are values.
    if len(filename) >= 53:
        log['timestamp'] = filename[39:53]
        log['cyclenumber'] = "Cycle " + str(filename[29:31])
    else:
        return

    # Needs a way to verify if this is a log file. If access to the file is not
    # allowed, cancel function.
    rawfile = safeOpen(filelocation, 'r')
    if not rawfile:
        return

    for i, rawline in enumerate(rawfile):
        line = rawline.replace('\n', '')

        # We expect to only see lines of 28 characters or more.
        if len(line) < 28:
            break

        # Check to see if the quantity is real or a duplicate. If real, get the
        # code and quantity and add it to the log dictionary
        if str(line[17:20]) == "QUA":
            cellcode = line[12:16]
            rawquantity = line[22:29]
            # Strip the string of all leading zeroes #
            quantity = rawquantity.lstrip("0")

            # If the cellcode currently exists in the log, add it to the
            # existing number. Otherwise make a new entry in the log dictionary
            # with key "cellcode" and value "quantity".
            if cellcode in log:
                log[cellcode] = log[cellcode] + quantity
            else:
                log[cellcode] = quantity

    return sorted(log.items())


###############################################################################
#   cell()                                                                    #
#                                                                             #
#   This function returns a cell position as a string. The default cell       #
#   returned with an input of (0,0) is "A4".                                  #
###############################################################################
def cell(column, row):

    # This function only works when column is less than 676. Above that... :(
    if column <= 25:
        columnvalue = string.ascii_uppercase[column]
    else:
        firstcharacter = int(math.floor(column / 26))-1
        secondcharacter = int(column % 26)
        columnvalue = string.ascii_uppercase[firstcharacter]+string.ascii_uppercase[secondcharacter]

    rowvalue = str(row + 4)
    return str(columnvalue + rowvalue)


###############################################################################
#   programCreateExcelFile()                                                  #
#                                                                             #
#   This function returns a Workbook object that has a sheet for every docket #
#   listed in the docket template file. A single Workbook is returned         #
###############################################################################
def programCreateExcelFile(docketrequirements, cellcounts):

    workbook = openpyxl.Workbook()
    infosheet = workbook.active

    infosheet.title = "Read Me"
    infosheet['B2'] = "This workbook was automatically generated."
    infosheet['B3'] = "It is advised that you copy and paste from this workbook into the proper files, instead of using this as a base."
    infosheet['B4'] = "The program that was used to create this file cannot update existing files."

    # Create a worksheet for every docket listed in the docket requirements
    # template.
    for docket in docketrequirements:
        worksheet = workbook.create_sheet()
        worksheet.title = docket.name
        worksheet['A1'] = docket.name
        worksheet['A3'] = "Cell Code"

        # These determine the location of the totals rows later
        maxwidth = len(cellcounts)+1
        maxheight = len(docket.requirements)+1

        # This function builds worksheets row by row from left to right.
        # It checks the current job docket for the codes that should be in the
        # cycle files
        for r, cellcode in enumerate(docket.requirements):

            # Set the code that it is looking for in the A column
            worksheet[cell(0, r)] = cellcode

            # Now check through each cyclefile to see if that code exists
            for c, cyclefile in enumerate(cellcounts):

                # Add 1 to column otherwise all cells will appear shifted left one.
                c += 1

                # Place the name of the cyclefile on row 3. This is done every
                # time the program goes down a row but that's okay.
                worksheet[cell(c, -1)] = cyclefile['timestamp']
                worksheet[cell(c, -2)] = cyclefile['cyclenumber']

                # If the cellcode exists in the current cyclefile, add the
                # quantity into the proper cell.
                if cellcode in cyclefile:
                    worksheet[cell(c, r)] = int(cyclefile[cellcode])
                    # This is to separate the known from the unknown.
                    cyclefile[cellcode] = "PARSED"
                else:
                    # If it doesn't exist in the current cyclefile, enter 0
                    worksheet[cell(c, r)] = 0

            # While the function is still iterating left to right, add a totals
            # cell on the right to show the total for that row.
            worksheet[cell(maxwidth, r)] = "=SUM(" + cell(1, r) + ":" + cell(maxwidth-1, r) + ")"

        # After the program is done adding all rows (it finished adding all
        # cellcodes from the docketrequirements file), this adds a totals row
        # on the bottom to show the totals from each cycle file.
        for totalscolumn in range(0, maxwidth + 1):
            worksheet[cell(totalscolumn, maxheight)] = "=SUM(" + cell(totalscolumn, 0) + ":" + cell(totalscolumn, r) + ")"

        # Just so that the worksheet looks nicer
        worksheet[cell(0, maxheight)] = "Cycle Total"
        worksheet[cell(maxwidth, -1)] = "Total Items Mailed"

    # PROCESSING COMPLETE   #
    # UNKNOWNS PARSED BELOW #

    # Create a worksheet to host all unknown values. It will be used later #
    unknowns = workbook.create_sheet()
    unknowns.title = "Extras"
    unknowns['A1'] = "Unknown Cell Codes"
    unknowns['A3'] = "Cell Code"
    unknowns['B3'] = "Quantity"
    unknowns['C3'] = "Cycle number"
    unknowns['D3'] = "Cycle timestamp"
    # This is increased later when we add rows to the sheet
    unknownsheetrow = 0

    # Read the cycle files again to find all codes that were not captured by
    # the first sweep. We'll put them on the unknowns worksheet.
    for cyclefile in cellcounts:
        for cellcode in cyclefile:
            if not cyclefile[cellcode] == "PARSED" and not cellcode == 'timestamp' and not cellcode == 'cyclenumber':
                unknowns[cell(0, unknownsheetrow)] = cellcode
                unknowns[cell(1, unknownsheetrow)] = int(cyclefile[cellcode])
                unknowns[cell(2, unknownsheetrow)] = cyclefile['cyclenumber']
                unknowns[cell(3, unknownsheetrow)] = cyclefile['timestamp']
                unknownsheetrow += 1

    return workbook


###############################################################################
